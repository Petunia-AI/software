from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc
from app.database import get_db
from app.models.lead import Lead, LeadStage, LeadSource
from app.api.auth import get_current_user
from app.models.user import User
from app.schemas.lead import LeadCreate, LeadUpdate, LeadOut
from typing import List, Optional
from datetime import datetime, timezone
import uuid
import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Columnas exportadas / importadas
EXPORT_FIELDS = [
    "name", "email", "phone", "company", "position",
    "stage", "source", "qualification_score",
    "budget", "authority", "need", "timeline",
    "estimated_value", "notes", "tags",
    "last_contacted_at", "next_followup_at", "created_at",
]

IMPORT_FIELDS = [
    "name", "email", "phone", "company", "position",
    "stage", "source", "estimated_value", "notes",
    "budget", "authority", "need", "timeline",
]

VALID_STAGES  = {e.value for e in LeadStage}
VALID_SOURCES = {e.value for e in LeadSource}

router = APIRouter(prefix="/leads", tags=["leads"])


@router.get("", response_model=List[LeadOut])
async def list_leads(
    stage: Optional[str] = None,
    source: Optional[str] = None,
    min_score: Optional[float] = None,
    limit: int = 50,
    offset: int = 0,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    query = (
        select(Lead)
        .where(Lead.business_id == current_user.business_id, Lead.is_active == True)
        .order_by(desc(Lead.created_at))
        .limit(limit)
        .offset(offset)
    )
    if stage:
        query = query.where(Lead.stage == stage)
    if source:
        query = query.where(Lead.source == source)
    if min_score is not None:
        query = query.where(Lead.qualification_score >= min_score)
    result = await db.execute(query)
    return result.scalars().all()


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT  GET /leads/export?format=csv|xlsx
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/export")
async def export_leads(
    format: str = Query("csv", regex="^(csv|xlsx)$"),
    stage: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    query = (
        select(Lead)
        .where(Lead.business_id == current_user.business_id, Lead.is_active == True)
        .order_by(desc(Lead.created_at))
    )
    if stage:
        query = query.where(Lead.stage == stage)
    result = await db.execute(query)
    leads = result.scalars().all()

    headers_map = {
        "name": "Nombre", "email": "Email", "phone": "Teléfono",
        "company": "Empresa", "position": "Cargo", "stage": "Etapa",
        "source": "Fuente", "qualification_score": "Score BANT",
        "budget": "Presupuesto", "authority": "Autoridad",
        "need": "Necesidad", "timeline": "Timeline",
        "estimated_value": "Valor estimado", "notes": "Notas",
        "tags": "Etiquetas", "last_contacted_at": "Último contacto",
        "next_followup_at": "Próximo seguimiento", "created_at": "Creado",
    }

    def get_val(lead, field):
        v = getattr(lead, field, None)
        if v is None:
            return ""
        if isinstance(v, datetime):
            return v.strftime("%Y-%m-%d %H:%M")
        if isinstance(v, (list, dict)):
            return str(v)
        return v

    if format == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=EXPORT_FIELDS)
        writer.writerow({f: headers_map[f] for f in EXPORT_FIELDS})
        for lead in leads:
            writer.writerow({f: get_val(lead, f) for f in EXPORT_FIELDS})
        output.seek(0)
        filename = f"leads_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        return StreamingResponse(
            iter([output.getvalue().encode("utf-8-sig")]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    # xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    header_fill  = PatternFill("solid", fgColor="6D28D9")
    header_font  = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, field in enumerate(EXPORT_FIELDS, 1):
        cell = ws.cell(row=1, column=col_idx, value=headers_map[field])
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = 20

    for row_idx, lead in enumerate(leads, 2):
        for col_idx, field in enumerate(EXPORT_FIELDS, 1):
            ws.cell(row=row_idx, column=col_idx, value=get_val(lead, field))

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"leads_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─────────────────────────────────────────────────────────────────────────────
# IMPORT  POST /leads/import
# ─────────────────────────────────────────────────────────────────────────────
@router.post("/import")
async def import_leads(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    content = await file.read()
    filename = (file.filename or "").lower()

    rows: list[dict] = []

    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)

    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        raw_rows = list(ws.iter_rows(values_only=True))
        if not raw_rows:
            raise HTTPException(status_code=400, detail="Archivo vacío")
        # Primera fila = cabeceras
        headers = [str(h).strip().lower() if h else "" for h in raw_rows[0]]
        for row in raw_rows[1:]:
            rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
    else:
        raise HTTPException(status_code=400, detail="Formato no soportado. Usa CSV o XLSX.")

    # Normalizar cabeceras en español → campo interno
    ALIAS = {
        # Nombre
        "nombre": "name", "nombre completo": "name", "nombre_completo": "name",
        "full name": "name", "full_name": "name", "contact": "name", "contacto": "name",
        # Email
        "email": "email", "correo": "email", "correo electrónico": "email",
        "correo electronico": "email", "e-mail": "email", "mail": "email",
        # Teléfono
        "teléfono": "phone", "telefono": "phone", "celular": "phone",
        "tel": "phone", "whatsapp": "phone", "móvil": "phone", "movil": "phone",
        "phone": "phone", "phone number": "phone", "número": "phone", "numero": "phone",
        # Empresa
        "empresa": "company", "negocio": "company", "compañia": "company",
        "compañía": "company", "company": "company", "organización": "company",
        "organizacion": "company", "razón social": "company", "razon social": "company",
        # Cargo
        "cargo": "position", "puesto": "position", "posición": "position",
        "posicion": "position", "title": "position", "job title": "position",
        "rol": "position", "role": "position",
        # Etapa
        "etapa": "stage", "estado": "stage", "stage": "stage", "status": "stage",
        # Fuente
        "fuente": "source", "canal": "source", "source": "source", "origen": "source",
        # Score
        "score bant": "qualification_score", "score": "qualification_score",
        "calificación": "qualification_score", "calificacion": "qualification_score",
        "puntuación": "qualification_score", "puntuacion": "qualification_score",
        "qualification_score": "qualification_score",
        # Valor
        "valor estimado": "estimated_value", "valor": "estimated_value",
        "presupuesto estimado": "estimated_value", "monto": "estimated_value",
        "importe": "estimated_value", "precio": "estimated_value",
        "estimated_value": "estimated_value",
        # Notas
        "notas": "notes", "nota": "notes", "observaciones": "notes",
        "comentarios": "notes", "notes": "notes", "comments": "notes",
        # BANT
        "presupuesto": "budget", "budget": "budget",
        "autoridad": "authority", "authority": "authority",
        "necesidad": "need", "need": "need", "necesidades": "need",
        "timeline": "timeline", "plazo": "timeline", "tiempo": "timeline",
        "tiempo de compra": "timeline",
    }

    created = 0
    skipped = 0
    errors: list[str] = []

    # Detectar columnas originales y cuáles no se reconocieron
    original_headers = list(rows[0].keys()) if rows else []
    mapped_fields = set(IMPORT_FIELDS)
    unrecognized_cols = [
        h for h in original_headers
        if ALIAS.get(h.strip().lower(), h.strip().lower()) not in mapped_fields
    ]

    for i, raw in enumerate(rows, 2):
        # Normalizar claves
        row: dict = {}
        for k, v in raw.items():
            key = ALIAS.get(k.strip().lower(), k.strip().lower())
            row[key] = v

        # Saltar filas completamente vacías
        if not any(v for v in row.values()):
            skipped += 1
            continue

        # name es NOT NULL — usar fallback: email, teléfono o "Sin nombre"
        name = (
            row.get("name") or
            row.get("email") or
            row.get("phone") or
            f"Lead fila {i}"
        ).strip() or f"Lead fila {i}"

        stage  = row.get("stage", "new").strip().lower()
        source = row.get("source", "manual").strip().lower()
        if stage  not in VALID_STAGES:  stage  = "new"
        if source not in VALID_SOURCES: source = "manual"

        try:
            score = float(row.get("qualification_score") or 0)
        except (ValueError, TypeError):
            score = 0.0

        try:
            est_val = float(row.get("estimated_value") or 0) or None
        except (ValueError, TypeError):
            est_val = None

        lead = Lead(
            id=str(uuid.uuid4()),
            business_id=current_user.business_id,
            name=name,
            email=row.get("email") or None,
            phone=row.get("phone") or None,
            company=row.get("company") or None,
            position=row.get("position") or None,
            stage=stage,
            source=source,
            qualification_score=score,
            estimated_value=est_val,
            notes=row.get("notes") or None,
            budget=row.get("budget") or None,
            authority=row.get("authority") or None,
            need=row.get("need") or None,
            timeline=row.get("timeline") or None,
            tags=[],
            is_active=True,
        )
        db.add(lead)
        created += 1

    try:
        await db.commit()
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al guardar: {str(e)}")

    return {
        "ok": True,
        "created": created,
        "skipped": skipped,
        "errors": errors,
        "unrecognized_columns": unrecognized_cols,
        "message": (
            f"{created} leads importados correctamente"
            + (f". Columnas no reconocidas: {', '.join(unrecognized_cols)}" if unrecognized_cols else "")
        ),
    }



@router.get("/{lead_id}", response_model=LeadOut)
async def get_lead(
    lead_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Lead).where(Lead.id == lead_id, Lead.business_id == current_user.business_id)
    )
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead no encontrado")
    return lead


@router.post("", response_model=LeadOut, status_code=201)
async def create_lead(
    data: LeadCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    lead = Lead(id=str(uuid.uuid4()), business_id=current_user.business_id, **data.model_dump())
    db.add(lead)
    await db.commit()
    await db.refresh(lead)
    return lead


@router.patch("/{lead_id}", response_model=LeadOut)
async def update_lead(
    lead_id: str,
    data: LeadUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Lead).where(Lead.id == lead_id, Lead.business_id == current_user.business_id)
    )
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead no encontrado")
    for key, value in data.model_dump(exclude_none=True).items():
        setattr(lead, key, value)
    await db.commit()
    await db.refresh(lead)
    return lead


@router.delete("/{lead_id}", status_code=204)
async def delete_lead(
    lead_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Lead).where(Lead.id == lead_id, Lead.business_id == current_user.business_id)
    )
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead no encontrado")
    lead.is_active = False
    await db.commit()


@router.patch("/{lead_id}/stage", response_model=LeadOut)
async def update_lead_stage(
    lead_id: str,
    body: dict,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Cambio rápido de etapa desde Kanban o tabla."""
    stage = body.get("stage", "")
    if stage not in VALID_STAGES:
        raise HTTPException(status_code=400, detail=f"Etapa inválida: {stage}")
    result = await db.execute(
        select(Lead).where(Lead.id == lead_id, Lead.business_id == current_user.business_id)
    )
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead no encontrado")
    lead.stage = stage
    await db.commit()
    await db.refresh(lead)
    return lead
